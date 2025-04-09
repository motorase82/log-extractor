import sqlite3
import os
import json
import pandas as pd
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import easyocr
import regex
from difflib import get_close_matches
from PIL import Image
import re
import datetime
import numpy as np
app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
RESULTS_FOLDER = "results"
ALLOWED_EXTENSIONS = {"txt", "db"}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["RESULTS_FOLDER"] = RESULTS_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)

# --------------------------- Player Info Extraction Global Setup --------------------------- #
reader = easyocr.Reader(['en'])
known_player_names = []

BAD_USERNAME_KEYWORDS = ["chat", "apply", "info", "achievement", "commander", "details", "more", "not set"]

def is_valid_username(name):
    return (
        name and
        not re.search(r"\d", name) and
        len(name) >= 3 and
        not any(bad in name.lower() for bad in BAD_USERNAME_KEYWORDS)
    )

def correct_name_with_list(ocr_name):
    match = get_close_matches(ocr_name, known_player_names, n=1, cutoff=0.85)
    return match[0] if match else ocr_name
# --------------------------- Image Preprocessing --------------------------- #
def preprocess_image(image_path):
    image = Image.open(image_path).convert("L")  # Convert to 8-bit grayscale
    threshold = 160
    image = image.point(lambda x: 255 if x > threshold else 0)  # Binary threshold, but stays in "L" mode
    return image


def cleanup_old_results():
    """Delete old Excel/CSV files from the results folder."""
    for fname in os.listdir(RESULTS_FOLDER):
        if fname.endswith((".xlsx", ".csv")):
            os.remove(os.path.join(RESULTS_FOLDER, fname))

def generate_timestamped_filenames(base: str):
    """Generate timestamped filenames for output."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return (
        f"{base}_{timestamp}.xlsx",
        f"{base}_{timestamp}.csv"
    )
def cleanup_old_results(days=1):
    cutoff = datetime.datetime.now() - datetime.timedelta(days=days)
    for fname in os.listdir(RESULTS_FOLDER):
        fpath = os.path.join(RESULTS_FOLDER, fname)
        if os.path.isfile(fpath):
            mtime = datetime.datetime.fromtimestamp(os.path.getmtime(fpath))
            if mtime < cutoff:
                os.remove(fpath)
                print(f"🗑️ Deleted old file: {fname}")
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# --------------------------- Zone Passes Count Extraction --------------------------- #
def extract_zone_passes(file_path):
    room_id = "custom_116351288000531_1741730818"
    conn = sqlite3.connect(file_path)
    cursor = conn.cursor()
    cursor.execute("SELECT SenderUid, Msg FROM chat WHERE RoomId = ?", (room_id,))
    messages = cursor.fetchall()

    zone_data = {}
    for uid, msg in messages:
        if msg.isdigit():
            value = int(msg)
            if uid not in zone_data or value > zone_data[uid]:
                zone_data[uid] = value

    try:
        user_info_df = pd.read_sql_query("SELECT Uid, UserName FROM chatUserInfo", conn)
    except:
        user_info_df = pd.DataFrame(columns=["Uid", "UserName"])

    df = pd.DataFrame([{"UID": uid, "Zone Passes": val} for uid, val in zone_data.items()])
    df = df.merge(user_info_df, left_on="UID", right_on="Uid", how="left")
    df.drop(columns=["Uid"], inplace=True)
    df.rename(columns={"UserName": "Name"}, inplace=True)
    df = df[["Name", "UID", "Zone Passes"]].sort_values(by="Zone Passes", ascending=False)

    excel_path = os.path.join(app.config["RESULTS_FOLDER"], "zone-passes-count_data.xlsx")
    csv_path = os.path.join(app.config["RESULTS_FOLDER"], "zone-passes-count_data.csv")
    df.to_excel(excel_path, index=False)
    df.to_csv(csv_path, index=False)

    print("✅ Saved Zone Passes to:", excel_path)
    conn.close()
    return df.to_dict(orient="records")

# --------------------------- Alliance Duel Points Extraction --------------------------- #
def extract_alliance_duel_points(file_path):
    try:
        conn = sqlite3.connect(file_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT Contents FROM mail
            WHERE ChannelId = 'system' AND Title = '361000' AND SubTitle = '361044'
            LIMIT 1;
        """)
        row = cursor.fetchone()

        if not row:
            cursor.execute("""
                SELECT Contents FROM mail
                WHERE ChannelId = 'system' AND LENGTH(Contents) > 1000
                ORDER BY CreateTime DESC
                LIMIT 1;
            """)
            row = cursor.fetchone()

        if not row:
            print("⚠️ No relevant AD data found.")
            return []

        ad_data = json.loads(row[0])
        player_scores = ad_data.get("allPlayerScore", [])
        cursor.execute("SELECT JsonStr FROM chatUserInfo")
        chat_data = cursor.fetchall()
        conn.close()

        uid_to_rank = {}
        for (json_str,) in chat_data:
            try:
                info = json.loads(json_str)
                uid_to_rank[info.get("uid")] = info.get("rank")
            except json.JSONDecodeError:
                continue

    except Exception as e:
        print(f"❌ DB/JSON error: {e}")
        return []

    score_data = []
    for player in player_scores:
        uid = player.get("uid", "Unknown UID")
        alliance_rank = uid_to_rank.get(uid, "N/A")
        score_data.append({
            "Player Name": player.get("name", "Unknown"),
            "Points": player.get("score", 0),
            "UID": uid,
            "Rank": player.get("rank", "-"),
            "Alliance Rank": alliance_rank
        })

    df = pd.DataFrame(score_data)

    excel_path = os.path.join(app.config["RESULTS_FOLDER"], "alliance-duel-points_data.xlsx")
    csv_path = os.path.join(app.config["RESULTS_FOLDER"], "alliance-duel-points_data.csv")

    # --------- Excel Export with Styling --------- #
    wb = Workbook()
    ws = wb.active
    ws.title = "Alliance Duel Points Data"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="8db4e2", end_color="8db4e2", fill_type="solid")

    ws.append(df.columns.tolist())
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(row)

        # Rank cell coloring
        rank = row.Rank
        rank_cell = ws.cell(row=row_idx, column=4)
        if isinstance(rank, int):
            if 1 <= rank <= 10:
                rank_cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
            elif 11 <= rank <= 30:
                rank_cell.fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
            elif 31 <= rank <= 60:
                rank_cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
            elif 61 <= rank <= 90:
                rank_cell.fill = PatternFill(start_color="ffa500", end_color="ffa500", fill_type="solid")

        # Points cell coloring
        points = row.Points
        points_cell = ws.cell(row=row_idx, column=2)
        if isinstance(points, (int, float)) and points < 24000000:
            points_cell.fill = PatternFill(start_color="ff6961", end_color="ff6961", fill_type="solid")

    ws.auto_filter.ref = ws.dimensions

    for col_num, column_title in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_num)
        max_length = max(len(str(column_title)), *(len(str(cell.value)) for cell in ws[col_letter]))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_path)
    df.to_csv(csv_path, index=False)

    print("✅ Saved Alliance Duel Points to:", excel_path)
    return df.to_dict(orient="records")
    
    
# --------------------------- Player Info Extraction --------------------------- #
def extract_combined_ocr_player_info(text):
    APC_SKINS = [
        "Plunder", "Demon", "Iron Fist", "Reindeer Sleigh", "Bigfoot",
        "Forerunner Motorcycle", "Daybreak", "Silver Phantom", "Pumpkin Phantom",
        "Polestar Express", "Party Yacht", "Duck Home", "Steam Tank", "Lunar Rover",
        "New Order", "Heavy Armor", "Desert Taxi", "Pumpkin Spiderling",
        "Run, Turkey! Run!", "Christmas Parade", "Springly Dragon Boat",
        "Armed Chariot", "None"
    ]

    STOPWORDS = {"lite", "lv", "lv.", "lv:", "chat", "apply", "details", "info", "not set", ""}

    player_data = {
        "Username": "",
        "ID": "",
        "CP": "",
        "Server": "",
        "Alliance": "",
        "APC": "",
        "Fortress Level": "",
        "Total Battles": "",
        "Battle Victories": "",
        "Units Defeated (Enemies)": "",
        "Units Defeated (Yours)": "",
        "Units Treated (Yours)": "",
        "Zombies Defeated": ""
    }

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    full_text = " ".join(lines)

    # --- Label-value block parsing from messy OCR ---
    label_map = {
        "fortress": "Fortress Level",
        "alliance": "Alliance",
        "apc": "APC"
    }

    i = 0
    while i < len(lines):
        line = lines[i].strip().lower()

        matched_label = None
        for key in label_map.keys():
            if get_close_matches(line, [key], n=1, cutoff=0.75):
                matched_label = key
                break

        if matched_label:
            field = label_map[matched_label]
            value_lines = []

            for j in range(1, 4):
                if i + j < len(lines):
                    val = lines[i + j].strip()
                    if val.lower() in label_map or val.lower() in STOPWORDS:
                        continue
                    value_lines.append(val)

            value = " ".join(value_lines).strip()
            if field == "APC":
                # Look for up to 4 lines after "APC"
                candidates = []
                k = i + 1
                lines_checked = 0

                while k < len(lines) and len(candidates) < 4 and lines_checked < 6:
                    candidate = lines[k].strip()
                    if (
                        not candidate
                        or re.fullmatch(r"\d{1,2}", candidate)
                        or "ctrl" in candidate.lower()
                        or "[" in candidate
                        or "lv" in candidate.lower()
                        or candidate.lower() in STOPWORDS
                    ):
                        k += 1
                        lines_checked += 1
                        continue
                    candidates.append(candidate)
                    k += 1
                    lines_checked += 1

                # Try fuzzy matching 1, 2, or 3 word combinations
                matched_apc = None
                for n in range(3, 0, -1):
                    for m in range(len(candidates) - n + 1):
                        combo = " ".join(candidates[m:m + n])
                        match = get_close_matches(combo.lower(), [skin.lower() for skin in APC_SKINS], n=1, cutoff=0.6)
                        if match:
                            matched_apc = next(s for s in APC_SKINS if s.lower() == match[0])
                            break
                    if matched_apc:
                        break

                player_data["APC"] = matched_apc if matched_apc else candidates[0] if candidates else ""

            elif field == "Fortress Level":
                match = re.search(r"(\d{1,2})", value)
                if match:
                    player_data["Fortress Level"] = match.group(1)
                else:
                    player_data["Fortress Level"] = value

            else:
                player_data[field] = value

            i += len(value_lines)

        i += 1  # ← this always runs to move the loop forward


# --- Extract ID and Server ---
    match = re.search(r"\bID[:：]?\s*([A-Z0-9]{5,})", full_text, re.IGNORECASE)
    if match:
        player_data["ID"] = match.group(1).upper()

    match = re.search(r"#(\d{2,4})", full_text)
    if match:
        player_data["Server"] = match.group(1)

# --- Username and CP ---
    for i, line in enumerate(lines):
        cp_match = re.search(r"\b\d{1,3}(?:,\d{3}){2,}\b", line)
        if cp_match:
            cp = cp_match.group(0)
            player_data["CP"] = cp
            possible_name = line.replace(cp, "").strip()
            if is_valid_username(possible_name):
                player_data["Username"] = correct_name_with_list(possible_name)
                break
            if i > 0 and is_valid_username(lines[i - 1].strip()):
                player_data["Username"] = correct_name_with_list(lines[i - 1].strip())
                break
            if i + 1 < len(lines) and is_valid_username(lines[i + 1].strip()):
                player_data["Username"] = correct_name_with_list(lines[i + 1].strip())
                break

# --- Extract Battle Stats ---
    stat_map = {
        "Total Battles": r"Total Battles\s+([\d,]+)",
        "Battle Victories": r"Battle Victories\s+([\d,]+)",
        "Units Defeated (Enemies)": r"Units Defeated\s*\(Enemies\)\s+([\d,]+)",
        "Units Defeated (Yours)": r"Units Defeated\s*\(Yours\)\s+([\d,]+)",
        "Units Treated (Yours)": r"Units Treated\s*\(Yours\)\s+([\d,]+)",
        "Zombies Defeated": r"Zombies Defeated\s+([\d,]+)"
    }

    for field, pattern in stat_map.items():
        match = re.search(pattern, full_text, re.IGNORECASE)
        if match:
            player_data[field] = match.group(1)

    return player_data

# --------------------------- Upload & Download Routes --------------------------- #
@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files.get("file")
    extraction_type = request.form.get("extraction_type")

    print("⚙️ Received upload request.")
    if file:
        print("Filename:", file.filename)
        print("Allowed?", allowed_file(file.filename))
    else:
        print("❌ No file received.")

    if not file or not allowed_file(file.filename):
        return jsonify({"error": "Invalid or missing file"}), 400

    if not extraction_type:
        return jsonify({"error": "Missing extraction_type"}), 400

    file_path = os.path.join(app.config["UPLOAD_FOLDER"], secure_filename(file.filename))
    file.save(file_path)

    print("📦 File saved to:", file_path)
    print("📁 File exists?", os.path.exists(file_path))
    print("📤 Extraction type:", extraction_type)

    if extraction_type == "alliance-duel-points":
        extracted_data = extract_alliance_duel_points(file_path)
    elif extraction_type == "zone-passes-count":
        extracted_data = extract_zone_passes(file_path)
    else:
        return jsonify({"error": "Unknown extraction_type"}), 400
    # Timestamp filenames
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"player-info_{timestamp}.xlsx"
    csv_filename = f"player-info_{timestamp}.csv"

    excel_path = os.path.join(app.config["RESULTS_FOLDER"], excel_filename)
    csv_path = os.path.join(app.config["RESULTS_FOLDER"], csv_filename)

    df.to_excel(excel_path, index=False)
    df.to_csv(csv_path, index=False)
    return jsonify({
        "message": f"Extracted info for {len(players_data)} players!",
        "player_count": len(players_data),
        "data": players_data,  # 🧪 for preview purposes
        "excel_file": f"/download-excel/{os.path.basename(excel_path)}",
        "csv_file": f"/download-csv/{os.path.basename(csv_path)}"
    })

@app.route("/download-excel/all-player-info")
def download_all_excel():
    files = [f for f in os.listdir(app.config["RESULTS_FOLDER"])
             if f.startswith("all-player-info_") and f.endswith(".xlsx")]
    if not files:
        return "No Excel file found.", 404
    latest = max(files, key=lambda f: os.path.getctime(os.path.join(app.config["RESULTS_FOLDER"], f)))
    return send_file(os.path.join(app.config["RESULTS_FOLDER"], latest), as_attachment=True)

@app.route("/download-csv/all-player-info")
def download_all_csv():
    files = [f for f in os.listdir(app.config["RESULTS_FOLDER"])
             if f.startswith("all-player-info_") and f.endswith(".csv")]
    if not files:
        return "No CSV file found.", 404
    latest = max(files, key=lambda f: os.path.getctime(os.path.join(app.config["RESULTS_FOLDER"], f)))
    return send_file(os.path.join(app.config["RESULTS_FOLDER"], latest), as_attachment=True)
# --------------------------- Player Info Extraction Upload--------------------------- #
@app.route("/player-info", methods=["GET", "POST"])
def player_info():
    if request.method == "GET":
        return render_template("player-info.html")

    uploaded_files = request.files.getlist("screenshots")

    if not uploaded_files:
        return jsonify({"error": "No screenshots uploaded."}), 400

    print(f"📥 Received {len(uploaded_files)} files.")

    saved_texts = []
    for file in uploaded_files:
        if file and file.filename:
            filename = secure_filename(file.filename)
            path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(path)
            text = "\n".join(reader.readtext(path, detail=0))
            print(f"📄 OCR from {filename}:\n{text}\n")
            saved_texts.append((filename, path, text))

    # Grouping by player ID
    player_groups = {}
    unknown_counter = 1

    for filename, path, text in saved_texts:
        match = re.search(r"\bID[:：]?\s*([A-Z0-9]{5,})", text, re.IGNORECASE)
        if match:
            player_id = match.group(1).upper()
        else:
            player_id = f"UNKNOWN_{unknown_counter}"
            unknown_counter += 1
        player_groups.setdefault(player_id, []).append(text)

    # Process each player group
    players_data = []
    for player_id, text_blocks in player_groups.items():
        combined_text = "\n".join(text_blocks)
        player_data = extract_combined_ocr_player_info(combined_text)
        players_data.append(player_data)

    df = pd.DataFrame(players_data)

    # Save results
    cleanup_old_results()
    excel_name, csv_name = generate_timestamped_filenames("all-player-info")
    excel_path = os.path.join(app.config["RESULTS_FOLDER"], excel_name)
    csv_path = os.path.join(app.config["RESULTS_FOLDER"], csv_name)
    df.to_excel(excel_path, index=False)
    df.to_csv(csv_path, index=False)

    print(f"✅ Saved data for {len(players_data)} players to {excel_name}")

    return jsonify({
        "message": f"Extracted info for {len(players_data)} players!",
        "player_count": len(players_data),
        "data": players_data,
        "excel_file": f"/download-excel/{excel_name}",
        "csv_file": f"/download-csv/{csv_name}"
    })


# --------------------------- Page Routes --------------------------- #
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/alliance-duel-points", methods=["GET", "POST"])
def alliance_duel_points():
    return render_template("alliance-duel-points.html", script="alliance_duel_points_script.js")

@app.route("/zone-passes-count", methods=["GET", "POST"])
def zone_passes_count():
    return render_template("zone-passes-count.html", script="zone_passes_count_script.js")

@app.route("/player-info-page")
def player_info_page():
    return render_template("player-info.html")

if __name__ == "__main__":
    app.run(debug=True)

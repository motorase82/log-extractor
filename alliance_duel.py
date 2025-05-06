# alliance_duel.py

import os
import sqlite3
import pandas as pd
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extract_alliance_duel_points(file_path, results_folder):
    try:
        conn = sqlite3.connect(file_path)
        cursor = conn.cursor()

        # Try to find Alliance Duel related mails
        cursor.execute("SELECT DISTINCT Title, SubTitle, LENGTH(Contents) FROM mail WHERE ChannelId = 'system'")
        print("🗃️ Available Titles/SubTitles in mail table:")
        for t, st, length in cursor.fetchall():
            print(f"  → Title: {t}, SubTitle: {st}, Length: {length}")

        cursor.execute("""
            SELECT Contents FROM mail
            WHERE ChannelId = 'system' AND Title = '361000' AND SubTitle = '361044'
            LIMIT 1;
        """)
        row = cursor.fetchone()

        if not row:
            print("🔍 No 361000/361044 entry. Trying fallback...")
            cursor.execute("""
                SELECT Contents FROM mail
                WHERE ChannelId = 'system' AND LENGTH(Contents) > 1000
                ORDER BY CreateTime DESC
                LIMIT 5;
            """)
            rows = cursor.fetchall()
        else:
            rows = [row]

        player_scores = []
        timestamp_ms = None
        for i, r in enumerate(rows):
            try:
                ad_data = json.loads(r[0])
            except json.JSONDecodeError as e:
                print(f"❌ JSON parse failed on row {i+1}:", e)
                continue

            for key in ["allPlayerScore", "topPlayerInfos", "rankInfo"]:
                if key in ad_data:
                    player_scores = ad_data[key]
                    timestamp_ms = ad_data.get("lastTime")
                    break

            if player_scores:
                break

        cursor.execute("SELECT JsonStr FROM chatUserInfo")
        chat_data = cursor.fetchall()
        conn.close()

        uid_to_rank = {}
        for (json_str,) in chat_data:
            try:
                info = json.loads(json_str)
                uid_to_rank[info.get("uid")] = info.get("rank")
            except json.JSONDecodeError:
                continue

        score_data = []
        timestamp = datetime.datetime.fromtimestamp(timestamp_ms / 1000).strftime("%Y-%m-%d %H:%M:%S") if timestamp_ms else "N/A"

        for player in player_scores:
            uid = player.get("uid", "Unknown UID")
            score_data.append({
                "Player Name": player.get("name", "Unknown"),
                "Points": player.get("score", 0),
                "UID": uid,
                "Rank": player.get("rank", "-"),
                "Alliance Rank": uid_to_rank.get(uid, "N/A"),
                "Timestamp": timestamp
            })

    except Exception as e:
        print(f"❌ DB/JSON error: {e}")
        score_data = []

    df = pd.DataFrame(score_data, columns=["Player Name", "Points", "UID", "Rank", "Alliance Rank", "Timestamp"])

    excel_path = os.path.join(results_folder, "alliance-duel-points_data.xlsx")
    csv_path = os.path.join(results_folder, "alliance-duel-points_data.csv")

    df.to_csv(csv_path, index=False)
    print("✅ Saved Alliance Duel Points CSV:", csv_path)

    if not df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Alliance Duel Points Data"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")

        for col_num, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Rank coloring
                if df.columns[col_idx - 1] == "Rank" and isinstance(value, int):
                    if value <= 10:
                        cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
                    elif value <= 30:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value <= 60:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    elif value <= 90:
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

                # Points warning color
                if df.columns[col_idx - 1] == "Points" and isinstance(value, (int, float)) and value < 24000000:
                    cell.fill = PatternFill(start_color="FF6961", end_color="FF6961", fill_type="solid")

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        ws.auto_filter.ref = ws.dimensions
        wb.save(excel_path)
        print("✅ Saved Alliance Duel Points Excel:", excel_path)
        print("✅ Saved Alliance Duel Points Excel:", excel_path)

    return df.to_dict(orient="records")
